import openpyxl
from rit_lib import *


    ################################
    #                              #
    #    CLASS DEFINITION          #
    #                              #
    ################################
    
class Info(struct):
    _slots = ( (str,"Name"),\
               (str,"Class"),\
               (str,"Link"),\
               (float,"Distance"),\
               (int,"NumOfTables"),\
               (int,"Price"),\
               (str,"Pizza?"),\
               (str,"Burgers?"),\
               (str,"Sandwitches?") )

    
    ################################
    #                              #
    # CONSTRUCTING LIST OF OBJECTS #
    #                              #
    ################################
    
def constructInfoObjects():
    List=[]
    x=openpyxl.load_workbook("C:\\Users\\Lord Daniel\\Desktop\\Uni Stuff\\Fall 2015-2016\\Exploration of place and space\\MAP PROJECT 2\\Info\\Pool_info.xlsx")
    s=x.get_sheet_by_name("Sheet1")
    for i in range(1,31):
        Obj = Info( s.cell(row=i+1, column=2).value,\
                    s.cell(row=i+1, column=3).value,\
                    s.cell(row=i+1, column=4).value,\
                    float(s.cell(row=i+1, column=5).value),\
                    s.cell(row=i+1, column=6).value,\
                    s.cell(row=i+1, column=7).value,\
                    s.cell(row=i+1, column=8).value,\
                    s.cell(row=i+1, column=9).value,\
                    s.cell(row=i+1, column=10).value )
        List += [ Obj ]
    return List

ObjList=constructInfoObjects()

for i in range(len(ObjList)):
    print("destination"+str(i+1)+" = "+'"'+ObjList[i].Name+'"')
    print("def setDestination"+str(i+1)+"():")
    print("    global DestinationLink")
    print("    DestinationLink = "+'"'+ObjList[i].Link+'"')
    print("    D = destinationLabel['text']")
    print("    D = destination"+str(i+1))
    print("    destinationLabel['text'] = D")
    print('destination_menu.add_command(label = "'+ObjList[i].Name+'", command = setDestination'+str(i+1)+')')
    print()
    
        
