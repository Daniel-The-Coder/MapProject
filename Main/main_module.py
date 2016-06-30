
__author__ = 'Lord Daniel'

from rit_lib import * # for classes
from tkinter import * # for GUI
import math
from PIL import Image, ImageDraw, ImageTk
from tkinter.font import Font
import openpyxl
import webbrowser
import time, datetime
import pprint
import threading
import smtplib


    ################################
    #                              #
    #    CLASS DEFINITION          #
    #                              #
    ################################
    
class Info(struct):
    _slots = ( (int,"Num"),\
               (str,"Name"),\
               (str,"Class"),\
               (str,"Link"),\
               (float,"Distance"),\
               (int,"NumOfTables"),\
               (int,"Price"),\
               (str,"Pizza"),\
               (str,"Burgers"),\
               (str,"Sandwiches") )

    
    ################################
    #                              #
    # CONSTRUCTING LIST OF OBJECTS #
    #                              #
    ################################
    
def constructInfoObjects():
    List=[]
    x=openpyxl.load_workbook("C:\\Users\\Lord Daniel\\Desktop\\Uni Stuff\\Fall 2015-2016\\Exploration of place and space\\MAP PROJECT 2\\Data\\Pool_info.xlsx")
    s=x.get_sheet_by_name("Sheet1")
    for i in range(1,31):
        Obj = Info( s.cell(row=i+1, column=1).value,\
                    s.cell(row=i+1, column=2).value,\
                    s.cell(row=i+1, column=3).value,\
                    s.cell(row=i+1, column=4).value,\
                    float(s.cell(row=i+1, column=5).value),\
                    s.cell(row=i+1, column=6).value,\
                    s.cell(row=i+1, column=7).value,\
                    s.cell(row=i+1, column=8).value,\
                    s.cell(row=i+1, column=9).value,\
                    s.cell(row=i+1, column=10).value )
        List += [ Obj ]
    return List

InfoObjList = constructInfoObjects()

#pp = pprint.PrettyPrinter()
#pp.pprint(InfoObjList)#
               
DestinationLink=""

##########################################
#                                        #
#                                        #
#    ################################    #
#    #                              #    #
#    #         GUI                  #    #
#    #                              #    #
#    ################################    #
#                                        #
#                                        #
##########################################


root = Tk()

root.title('8 Ball Pool Info Station')

root.configure(background="grey")

canvas = Canvas(root, width=root.winfo_screenwidth(), height=root.winfo_screenheight())#
canvas.pack()#
tk_img = ImageTk.PhotoImage(file = 'pool.jpg')
canvas.create_image(800,450, image=tk_img)

'''
    ################################
    #                              #
    #         TIME  LABEL          #
    #                              #
    ################################

dt = datetime.datetime.now()
DATE = "Date: "+ str(dt.month) + "/" + str(dt.day) + "/" + str(dt.year)
TIME = "Time: "+ str(dt.hour) + ":" + str(dt.minute)

timeLabel = Label(root, text=TIME)
timeLabel.config(font=('helvetica',10))
timeLabel.place(relx=0.5, rely=0.02)

dateLabel = Label(root, text=DATE)
dateLabel.config(font=('helvetica',10))
dateLabel.place(relx=0.5, rely=0.04)

def DateTime():
    time.sleep(1)
    dt = datetime.datetime.now()
    
    DATE = "Date: "+ str(dt.month) + "/" + str(dt.day) + "/" + str(dt.year)
    TIME = "Time: "+ str(dt.hour) + ":" + str(dt.minute) + ":" + str(dt.second)
    
    A = timeLabel['text']
    A = TIME
    timeLabel['text'] = A
    
    B = dateLabel['text']
    B = DATE
    dateLabel['text'] = B

threadObject = threading.Thread(target = DateTime)
threadObject.start()
'''
    ################################
    #                              #
    #         IMAGE LABEL          #
    #                              #
    ################################

MAP=PhotoImage(file="marked_map_resized.png")
imageLabel = Label(root, image=MAP)
imageLabel. place(relx=0.4, rely=0.04)


    ################################
    #                              #
    #    SWITCH MAP BUTTONS        #
    #                              #
    ################################

################### DUBAI MAP ###################
'''
def Disp(c):
    IMG = Image.open(c)
    Img = ImageTk.PhotoImage(IMG)
    imageLabel.configure(image=Img)
    imageLabel.image = Img
    time.sleep(0.5)

def DispDubaiMap():
    for ImageFile in ["BurjKhalifa1_resized.png", "BurjKhalifa2_resized.png",\
                  "BurjKhalifa3_resized.png", "BurjKhalifa4_resized.png",\
                  "BurjKhalifa5_resized.png" ]:
        Disp(ImageFile)
'''
def DispDubaiMap():
    IMG = Image.open("dxbMap_marked_resized.png")
    Img = ImageTk.PhotoImage(IMG)
    imageLabel.configure(image=Img)
    imageLabel.image = Img
    
b1 = Button(root, text=" Dubai Map", command=DispDubaiMap, anchor='w') 
b1.config(height=2, width=10) # button size
b1.config(font=('algerian', 20)) # font
b1.config(bg="black", fg="white", activebackground="red", activeforeground="black")
b1.config(highlightbackground="grey")
b1_window = canvas.create_window(200, 350, anchor='nw', window=b1)  

################### ROCHESTER MAP ###################

def DispRochesterMap():
    IMG = Image.open("marked_map_resized.png")
    Img = ImageTk.PhotoImage(IMG)
    imageLabel.configure(image=Img)
    imageLabel.image = Img

b2 = Button(root, text=" Rochester Map", command=DispRochesterMap, anchor='w') 
b2.config(height=2, width=16) # button size
b2.config(font=('algerian', 20)) # font
b2.config(bg="indigo", fg="yellow", activebackground="red", activeforeground="black")
b2.config(highlightbackground="grey")
b2_window = canvas.create_window(200, 460, anchor='nw', window=b2)



    ################################
    #                              #
    #     SHOW ROUTE BUTTON        #
    #                              #
    ################################

def ShowRoute():
    global DestinationLink
    if DestinationLink == "":
        pass
    else:
        webbrowser.open(DestinationLink)
        
b3 = Button(root, text=" Show Route", command=ShowRoute, anchor='w') 
b3.config(height=2, width=12) # button size
b3.config(font=('algerian', 20)) # font
b3.config(bg="red", fg="white", activebackground="black", activeforeground="white")
b3.config(highlightbackground="grey")
b3_window = canvas.create_window(150, 170, anchor='nw', window=b3)


    ################################
    #                              #
    #        RESET   BUTTON        #
    #                              #
    ################################
    
def ResetLabels():
    for Label in LabelsList:
        Lbl = Label['text']
        Lbl = "                                     "
        Label['text'] = Lbl
        
b4 = Button(root, text=" Reset", command=ResetLabels, anchor='w') 
b4.config(height=1, width=7) # button size
b4.config(font=('algerian', 15)) # font
b4.config(bg="white", fg="black", activebackground="white", activeforeground="white")
b4.config(highlightbackground="grey")
b4_window = canvas.create_window(150, 650, anchor='nw', window=b4)


    ################################
    #                              #
    #      DISPLAY ALL BUTTON      #
    #                              #
    ################################

def DisplayAll():
    L = len(InfoObjList)
    for i in range(L):
        txt = str(InfoObjList[i].Num)+" "+InfoObjList[i].Name
        Lbl = LabelsList[i]['text']
        Lbl = txt
        LabelsList[i]['text'] = Lbl
b5 = Button(root, text=" Display All", command=DisplayAll, anchor='w') 
b5.config(height=1, width=11) # button size
b5.config(font=('algerian', 15)) # font
b5.config(bg="white", fg="black", activebackground="white", activeforeground="white")
b5.config(highlightbackground="grey")
b5_window = canvas.create_window(300, 650, anchor='nw', window=b5)
    

    ################################
    #                              #
    #         TEXT LABELS          #
    #                              #
    ################################

destinationLabel = Label(root, text="No destination chosen", font=('helvetica',20))
destinationLabel.place(relx=0.1, rely=0.1)

label1 = Label(root, text="1                                 ")
label1.config(font=('helvetica',10))
label1.place(relx=0.77, rely=0.02)

label2 = Label(root, text="2                                 ")
label2.config(font=('helvetica',10))
label2.place(relx=0.77, rely=0.05)

label3 = Label(root, text="3                                 ")
label3.config(font=('helvetica',10))
label3.place(relx=0.77, rely=0.08)

label4 = Label(root, text="4                                 ")
label4.config(font=('helvetica',10))
label4.place(relx=0.77, rely=0.11)

label5 = Label(root, text="5                                 ")
label5.config(font=('helvetica',10))
label5.place(relx=0.77, rely=0.13999999999999999)

label6 = Label(root, text="6                                 ")
label6.config(font=('helvetica',10))
label6.place(relx=0.77, rely=0.16999999999999998)

label7 = Label(root, text="7                                 ")
label7.config(font=('helvetica',10))
label7.place(relx=0.77, rely=0.19999999999999998)

label8 = Label(root, text="8                                 ")
label8.config(font=('helvetica',10))
label8.place(relx=0.77, rely=0.22999999999999998)

label9 = Label(root, text="9                                 ")
label9.config(font=('helvetica',10))
label9.place(relx=0.77, rely=0.26)

label10 = Label(root, text="10                                 ")
label10.config(font=('helvetica',10))
label10.place(relx=0.77, rely=0.29000000000000004)

label11 = Label(root, text="11                                 ")
label11.config(font=('helvetica',10))
label11.place(relx=0.77, rely=0.32)

label12 = Label(root, text="12                                 ")
label12.config(font=('helvetica',10))
label12.place(relx=0.77, rely=0.35)

label13 = Label(root, text="13                                 ")
label13.config(font=('helvetica',10))
label13.place(relx=0.77, rely=0.38)

label14 = Label(root, text="14                                 ")
label14.config(font=('helvetica',10))
label14.place(relx=0.77, rely=0.41000000000000003)

label15 = Label(root, text="15                                 ")
label15.config(font=('helvetica',10))
label15.place(relx=0.77, rely=0.44)

label16 = Label(root, text="16                                 ")
label16.config(font=('helvetica',10))
label16.place(relx=0.77, rely=0.47)

label17 = Label(root, text="17                                 ")
label17.config(font=('helvetica',10))
label17.place(relx=0.77, rely=0.5)

label18 = Label(root, text="18                                 ")
label18.config(font=('helvetica',10))
label18.place(relx=0.77, rely=0.53)

label19 = Label(root, text="19                                 ")
label19.config(font=('helvetica',10))
label19.place(relx=0.77, rely=0.56)

label20 = Label(root, text="20                                 ")
label20.config(font=('helvetica',10))
label20.place(relx=0.77, rely=0.59)

label21 = Label(root, text="21                                 ")
label21.config(font=('helvetica',10))
label21.place(relx=0.77, rely=0.62)

label22 = Label(root, text="22                                 ")
label22.config(font=('helvetica',10))
label22.place(relx=0.77, rely=0.65)

label23 = Label(root, text="23                                 ")
label23.config(font=('helvetica',10))
label23.place(relx=0.77, rely=0.6799999999999999)

label24 = Label(root, text="24                                 ")
label24.config(font=('helvetica',10))
label24.place(relx=0.77, rely=0.71)

label25 = Label(root, text="25                                 ")
label25.config(font=('helvetica',10))
label25.place(relx=0.77, rely=0.74)

label26 = Label(root, text="26                                 ")
label26.config(font=('helvetica',10))
label26.place(relx=0.77, rely=0.77)

label27 = Label(root, text="27                                 ")
label27.config(font=('helvetica',10))
label27.place(relx=0.77, rely=0.8)

label28 = Label(root, text="28                                 ")
label28.config(font=('helvetica',10))
label28.place(relx=0.77, rely=0.83)

label29 = Label(root, text="29                                 ")
label29.config(font=('helvetica',10))
label29.place(relx=0.77, rely=0.86)

label30 = Label(root, text="30                                 ")
label30.config(font=('helvetica',10))
label30.place(relx=0.77, rely=0.89)

LabelsList = [ label1, label2, label3, label4, label5, label6, label7,\
               label8, label9, label10, label11, label12, label13, label14,\
               label15, label16, label17, label18, label19, label20, label21,\
               label22, label23, label24, label25, label26, label27, label28,\
               label29, label30 ]


    ################################
    #                              #
    #         MENU BAR             #
    #                              #
    ################################

menubar = Menu(root)


    ################################
    #                              #
    #      DESTINATION MENU        #
    #                              #
    ################################
destination_menu= Menu(menubar, tearoff=0)

destination1 = "The Reunion Inn"
def setDestination1():
    global DestinationLink
    DestinationLink = "https://www.google.com/maps/dir/Rochester+Institute+of+Technology,+1+Lomb+Memorial+Drive,+Rochester,+NY+14623/Reunion+Inn,+4565+Culver+Road,+Rochester,+NY+14622/@43.1576622,-77.7497912,11z/data=!3m1!4b1!4m14!4m13!1m5!1m1!1s0x89d14c603a147e55:0xbe7eb31ed5e22c03!2m2!1d-77.6756115!2d43.0832027!1m5!1m1!1s0x89d6c830cb12c76d:0xb7b5105cbcd1d406!2m2!1d-77.5448456!2d43.2322918!3e1"
    D = destinationLabel['text']
    D = destination1
    destinationLabel['text'] = D
destination_menu.add_command(label = "The Reunion Inn", command = setDestination1)

destination2 = "Windjammers Restaurant"
def setDestination2():
    global DestinationLink
    DestinationLink = "https://www.google.com/maps/dir/Rochester+Institute+of+Technology,+1+Lomb+Memorial+Drive,+Rochester,+NY+14623/Windjammers,+4695+Lake+Avenue,+Rochester,+NY+14612/@43.1701505,-77.7844535,11z/am=t/data=!3m1!4b1!4m14!4m13!1m5!1m1!1s0x89d14c603a147e55:0xbe7eb31ed5e22c03!2m2!1d-77.6756115!2d43.0832027!1m5!1m1!1s0x89d6b7765dc6137d:0x5e2ccef935f699f1!2m2!1d-77.6103504!2d43.256456!3e1"
    D = destinationLabel['text']
    D = destination2
    destinationLabel['text'] = D
destination_menu.add_command(label = "Windjammers Restaurant", command = setDestination2)

destination3 = "Rab's Woodshed"
def setDestination3():
    global DestinationLink
    DestinationLink = "https://www.google.com/maps/dir/Rochester+Institute+of+Technology,+1+Lomb+Memorial+Drive,+Rochester,+NY+14623/Rab's+Woodshed,+4440+Lake+Avenue,+Rochester,+NY+14612/@43.1673877,-77.7795061,11z/data=!3m1!4b1!4m14!4m13!1m5!1m1!1s0x89d14c603a147e55:0xbe7eb31ed5e22c03!2m2!1d-77.6756115!2d43.0832027!1m5!1m1!1s0x89d6b774a6b2e6e5:0x880b0a45caef44de!2m2!1d-77.6132208!2d43.2515545!3e1"
    D = destinationLabel['text']
    D = destination3
    destinationLabel['text'] = D
destination_menu.add_command(label = "Rab's Woodshed", command = setDestination3)

destination4 = "Sundowners Inn"
def setDestination4():
    global DestinationLink
    DestinationLink = "https://www.google.com/maps/dir/Rochester+Institute+of+Technology,+1+Lomb+Memorial+Drive,+Rochester,+NY+14623/Sundowners+Inn,+19+Saint+Johns+Park+%231,+Rochester,+NY+14612/@43.1632712,-77.7893459,11z/data=!3m1!4b1!4m14!4m13!1m5!1m1!1s0x89d14c603a147e55:0xbe7eb31ed5e22c03!2m2!1d-77.6756115!2d43.0832027!1m5!1m1!1s0x89d6b70fc3eddd6f:0x1156380a66790bf0!2m2!1d-77.6183037!2d43.2430365!3e1"
    D = destinationLabel['text']
    D = destination4
    destinationLabel['text'] = D
destination_menu.add_command(label = "Sundowners Inn", command = setDestination4)

destination5 = "Romig's Tavern"
def setDestination5():
    global DestinationLink
    DestinationLink = "https://www.google.com/maps/dir/Rochester+Institute+of+Technology,+1+Lomb+Memorial+Drive,+Rochester,+NY+14623/Romig's+Tavern,+18+Bennington+Drive,+Rochester,+NY+14616/@43.149014,-77.7194728,12z/data=!3m1!4b1!4m14!4m13!1m5!1m1!1s0x89d14c603a147e55:0xbe7eb31ed5e22c03!2m2!1d-77.6756115!2d43.0832027!1m5!1m1!1s0x89d6b6b864b47657:0xba43fa5727cc809!2m2!1d-77.6393843!2d43.2147229!3e1"
    D = destinationLabel['text']
    D = destination5
    destinationLabel['text'] = D
destination_menu.add_command(label = "Romig's Tavern", command = setDestination5)

destination6 = "Barnards Crossing"
def setDestination6():
    global DestinationLink
    DestinationLink = "https://www.google.com/maps/dir/Rochester+Institute+of+Technology,+1+Lomb+Memorial+Drive,+Rochester,+NY+14623/Romig's+Tavern,+18+Bennington+Drive,+Rochester,+NY+14616/@43.149014,-77.7194728,12z/data=!3m1!4b1!4m14!4m13!1m5!1m1!1s0x89d14c603a147e55:0xbe7eb31ed5e22c03!2m2!1d-77.6756115!2d43.0832027!1m5!1m1!1s0x89d6b6b864b47657:0xba43fa5727cc809!2m2!1d-77.6393843!2d43.2147229!3e1"
    D = destinationLabel['text']
    D = destination6
    destinationLabel['text'] = D
destination_menu.add_command(label = "Barnards Crossing", command = setDestination6)

destination7 = "California Brew Haus"
def setDestination7():
    global DestinationLink
    DestinationLink = "https://www.google.com/maps/dir/Rochester+Institute+of+Technology,+1+Lomb+Memorial+Drive,+Rochester,+NY+14623/California+Brew+Haus,+402+West+Ridge+Road,+Rochester,+NY+14615/@43.1418303,-77.7194728,12z/data=!3m1!4b1!4m14!4m13!1m5!1m1!1s0x89d14c603a147e55:0xbe7eb31ed5e22c03!2m2!1d-77.6756115!2d43.0832027!1m5!1m1!1s0x89d6b6a6bc26346f:0x52223e3847ef292!2m2!1d-77.6375078!2d43.2002363!3e1"
    D = destinationLabel['text']
    D = destination7
    destinationLabel['text'] = D
destination_menu.add_command(label = "California Brew Haus", command = setDestination7)

destination8 = "Bathtub Billy's"
def setDestination8():
    global DestinationLink
    DestinationLink = "https://www.google.com/maps/dir/Rochester+Institute+of+Technology,+1+Lomb+Memorial+Drive,+Rochester,+NY+14623/Bathtub+Billy's+Restaurant+%26+Sports+Bar,+630+W+Ridge+Rd+%23+10,+Rochester,+NY+14615/@43.1434614,-77.7194728,12z/data=!3m1!4b1!4m14!4m13!1m5!1m1!1s0x89d14c603a147e55:0xbe7eb31ed5e22c03!2m2!1d-77.6756115!2d43.0832027!1m5!1m1!1s0x89d6b427541c66fb:0x91be0a76215e363d!2m2!1d-77.643418!2d43.2035247!3e1"
    D = destinationLabel['text']
    D = destination8
    destinationLabel['text'] = D
destination_menu.add_command(label = "Bathtub Billy's", command = setDestination8)

destination9 = "Straight Home"
def setDestination9():
    global DestinationLink
    DestinationLink = "https://www.google.com/maps/dir/Rochester+Institute+of+Technology,+1+Lomb+Memorial+Drive,+Rochester,+NY+14623/Straight+Home+Inn+Pizzeria,+688+Lexington+Avenue,+Rochester,+NY+14613/@43.1312691,-77.7194728,12z/data=!3m1!4b1!4m14!4m13!1m5!1m1!1s0x89d14c603a147e55:0xbe7eb31ed5e22c03!2m2!1d-77.6756115!2d43.0832027!1m5!1m1!1s0x89d6b40b42a43819:0xe96c09820f6f7e0f!2m2!1d-77.6486145!2d43.1793781!3e1"
    D = destinationLabel['text']
    D = destination9
    destinationLabel['text'] = D
destination_menu.add_command(label = "Straight Home", command = setDestination9)

destination10 = "Scotland Yard"
def setDestination10():
    global DestinationLink
    DestinationLink = "https://www.google.com/maps/dir/Rochester+Institute+of+Technology,+1+Lomb+Memorial+Drive,+Rochester,+NY+14623/Scotland+Yard+Pub,+187+Saint+Paul+Street,+Rochester,+NY+14604/@43.1218122,-77.7154138,12z/data=!3m1!4b1!4m14!4m13!1m5!1m1!1s0x89d14c603a147e55:0xbe7eb31ed5e22c03!2m2!1d-77.6756115!2d43.0832027!1m5!1m1!1s0x89d6b453712c9423:0xb40c7370d782c06!2m2!1d-77.6105271!2d43.1604321!3e1"
    D = destinationLabel['text']
    D = destination10
    destinationLabel['text'] = D
destination_menu.add_command(label = "Scotland Yard", command = setDestination10)

destination11 = "Joey's"
def setDestination11():
    global DestinationLink
    DestinationLink = "https://www.google.com/maps/dir/Rochester+Institute+of+Technology,+1+Lomb+Memorial+Drive,+Rochester,+NY+14623/Scotland+Yard+Pub,+187+Saint+Paul+Street,+Rochester,+NY+14604/@43.1218122,-77.7154138,12z/data=!3m1!4b1!4m14!4m13!1m5!1m1!1s0x89d14c603a147e55:0xbe7eb31ed5e22c03!2m2!1d-77.6756115!2d43.0832027!1m5!1m1!1s0x89d6b453712c9423:0xb40c7370d782c06!2m2!1d-77.6105271!2d43.1604321!3e1"
    D = destinationLabel['text']
    D = destination11
    destinationLabel['text'] = D
destination_menu.add_command(label = "Joey's", command = setDestination11)

destination12 = "Masons on Alexander"
def setDestination12():
    global DestinationLink
    DestinationLink = "https://www.google.com/maps/dir/Rochester+Institute+of+Technology,+1+Lomb+Memorial+Drive,+Rochester,+NY+14623/Masons+On+Alexander,+Alexander+Street,+Rochester,+NY/@43.1201092,-77.7048118,12z/data=!3m1!4b1!4m14!4m13!1m5!1m1!1s0x89d14c603a147e55:0xbe7eb31ed5e22c03!2m2!1d-77.6756115!2d43.0832027!1m5!1m1!1s0x89d6b5a7b3eae1e1:0xa9003d0ecbc0565a!2m2!1d-77.5944581!2d43.155393!3e1"
    D = destinationLabel['text']
    D = destination12
    destinationLabel['text'] = D
destination_menu.add_command(label = "Masons on Alexander", command = setDestination12)

destination13 = "Marshall Street Bar & Grill"
def setDestination13():
    global DestinationLink
    DestinationLink = "https://www.google.com/maps/dir/Rochester+Institute+of+Technology,+1+Lomb+Memorial+Drive,+Rochester,+NY+14623/Marshall+Street+Bar+and+Grill,+Marshall+Street,+Rochester,+NY/@43.1167846,-77.7079535,12z/data=!3m1!4b1!4m14!4m13!1m5!1m1!1s0x89d14c603a147e55:0xbe7eb31ed5e22c03!2m2!1d-77.6756115!2d43.0832027!1m5!1m1!1s0x89d6b506955016fb:0x257d50c4432cf5ce!2m2!1d-77.6005586!2d43.1502051!3e1"
    D = destinationLabel['text']
    D = destination13
    destinationLabel['text'] = D
destination_menu.add_command(label = "Marshall Street Bar & Grill", command = setDestination13)

destination14 = "Dicky's"
def setDestination14():
    global DestinationLink
    DestinationLink = "https://www.google.com/maps/dir/Rochester+Institute+of+Technology,+1+Lomb+Memorial+Drive,+Rochester,+NY+14623/Dicky's,+791+Meigs+Street,+Rochester,+NY+14620/@43.1124899,-77.707586,12z/data=!3m1!4b1!4m14!4m13!1m5!1m1!1s0x89d14c603a147e55:0xbe7eb31ed5e22c03!2m2!1d-77.6756115!2d43.0832027!1m5!1m1!1s0x89d6b5197b43b5f7:0x20640ea55be86034!2m2!1d-77.6000053!2d43.1391862!3e1"
    D = destinationLabel['text']
    D = destination14
    destinationLabel['text'] = D
destination_menu.add_command(label = "Dicky's", command = setDestination14)

destination15 = "Monty's Krown"
def setDestination15():
    global DestinationLink
    DestinationLink = "https://www.google.com/maps/dir/Rochester+Institute+of+Technology,+1+Lomb+Memorial+Drive,+Rochester,+NY+14623/Monty's+Krown,+875+Monroe+Avenue,+Rochester,+NY+14620/@43.1159357,-77.7701603,11z/data=!3m1!4b1!4m14!4m13!1m5!1m1!1s0x89d14c603a147e55:0xbe7eb31ed5e22c03!2m2!1d-77.6756115!2d43.0832027!1m5!1m1!1s0x89d6b512b71e13eb:0xe57e9e36f5037760!2m2!1d-77.5853751!2d43.1415766!3e1"
    D = destinationLabel['text']
    D = destination15
    destinationLabel['text'] = D
destination_menu.add_command(label = "Monty's Krown", command = setDestination15)

destination16 = "Acme"
def setDestination16():
    global DestinationLink
    DestinationLink = "https://www.google.com/maps/dir/Rochester+Institute+of+Technology,+1+Lomb+Memorial+Drive,+Rochester,+NY+14623/Acme+Bar+%26+Pizza,+495+Monroe+Avenue,+Rochester,+NY+14607/@43.1159994,-77.7045825,12z/data=!3m1!4b1!4m14!4m13!1m5!1m1!1s0x89d14c603a147e55:0xbe7eb31ed5e22c03!2m2!1d-77.6756115!2d43.0832027!1m5!1m1!1s0x89d6b50f1b333529:0x89e07ca81da419a9!2m2!1d-77.5939706!2d43.1464673!3e1"
    D = destinationLabel['text']
    D = destination16
    destinationLabel['text'] = D
destination_menu.add_command(label = "Acme", command = setDestination16)

destination17 = "O'Callaghan's"
def setDestination17():
    global DestinationLink
    DestinationLink = "https://www.google.com/maps/dir/Rochester+Institute+of+Technology,+1+Lomb+Memorial+Drive,+Rochester,+NY+14623/O'Callaghan's+Pub,+470+Monroe+Avenue,+Rochester,+NY+14607/@43.1159994,-77.7047192,12z/data=!3m1!4b1!4m14!4m13!1m5!1m1!1s0x89d14c603a147e55:0xbe7eb31ed5e22c03!2m2!1d-77.6756115!2d43.0832027!1m5!1m1!1s0x89d6b50f10d12099:0x232f476b9d923405!2m2!1d-77.5940899!2d43.1470076!3e1"
    D = destinationLabel['text']
    D = destination17
    destinationLabel['text'] = D
destination_menu.add_command(label = "O'Callaghan's", command = setDestination17)

destination18 = "Angry Duck"
def setDestination18():
    global DestinationLink
    DestinationLink = "https://www.google.com/maps/dir/Rochester+Institute+of+Technology,+1+Lomb+Memorial+Drive,+Rochester,+NY+14623/Angry+Duck,+510+Monroe+Avenue,+Rochester,+NY+14607/@43.1159994,-77.7043677,12z/data=!3m1!4b1!4m14!4m13!1m5!1m1!1s0x89d14c603a147e55:0xbe7eb31ed5e22c03!2m2!1d-77.6756115!2d43.0832027!1m5!1m1!1s0x89d6b50efd476363:0xbc0db21cbae9d53b!2m2!1d-77.5933869!2d43.1464999!3e1"
    D = destinationLabel['text']
    D = destination18
    destinationLabel['text'] = D
destination_menu.add_command(label = "Angry Duck", command = setDestination18)

destination19 = "Dr's Inn"
def setDestination19():
    global DestinationLink
    DestinationLink = "https://www.google.com/maps/dir/Rochester+Institute+of+Technology,+1+Lomb+Memorial+Drive,+Rochester,+NY+14623/Dr's+Inn,+1743+East+Avenue,+Rochester,+NY+14610/@43.1200454,-77.7564508,11z/data=!3m1!4b1!4m14!4m13!1m5!1m1!1s0x89d14c603a147e55:0xbe7eb31ed5e22c03!2m2!1d-77.6756115!2d43.0832027!1m5!1m1!1s0x89d6ca816cfada6d:0x98b622730d5a0559!2m2!1d-77.5579354!2d43.1457754!3e1"
    D = destinationLabel['text']
    D = destination19
    destinationLabel['text'] =D
destination_menu.add_command(label = "Dr's Inn", command = setDestination19)

destination20 = "Colters"
def setDestination20():
    global DestinationLink
    DestinationLink = "https://www.google.com/maps/dir/Rochester+Institute+of+Technology,+1+Lomb+Memorial+Drive,+Rochester,+NY+14623/Colter's,+262+Winton+Road+North,+Rochester,+NY+14610/@43.1200454,-77.7536759,11z/data=!3m1!4b1!4m14!4m13!1m5!1m1!1s0x89d14c603a147e55:0xbe7eb31ed5e22c03!2m2!1d-77.6756115!2d43.0832027!1m5!1m1!1s0x89d6ca7d1a461e11:0x1174e52560556de4!2m2!1d-77.5522573!2d43.1505141!3e1"
    D = destinationLabel['text']
    D = destination20
    destinationLabel['text'] = D
destination_menu.add_command(label = "Colters", command = setDestination20)

destination21 = "Jack Ryan's"
def setDestination21():
    global DestinationLink
    DestinationLink = "https://www.google.com/maps/dir/Rochester+Institute+of+Technology,+1+Lomb+Memorial+Drive,+Rochester,+NY+14623/Jack+Ryan's+Tavern,+825+Atlantic+Avenue,+Rochester,+NY+14609/@43.1200454,-77.7581577,11z/data=!3m1!4b1!4m14!4m13!1m5!1m1!1s0x89d14c603a147e55:0xbe7eb31ed5e22c03!2m2!1d-77.6756115!2d43.0832027!1m5!1m1!1s0x89d6ca78886be80b:0xcbf1663638a9010b!2m2!1d-77.5612209!2d43.156424!3e1"
    D = destinationLabel['text']
    D = destination21
    destinationLabel['text'] = D
destination_menu.add_command(label = "Jack Ryan's", command = setDestination21)

destination22 = "Main Place"
def setDestination22():
    global DestinationLink
    DestinationLink = "https://www.google.com/maps/dir/Rochester+Institute+of+Technology,+1+Lomb+Memorial+Drive,+Rochester,+NY+14623/Main+Place+Tavern,+East+Main+Street,+Rochester,+NY/@43.1225362,-77.7576212,11z/data=!3m1!4b1!4m14!4m13!1m5!1m1!1s0x89d14c603a147e55:0xbe7eb31ed5e22c03!2m2!1d-77.6756115!2d43.0832027!1m5!1m1!1s0x89d6ca779186cae3:0xed529fbd5b89a70e!2m2!1d-77.560166!2d43.159668!3e1"
    D = destinationLabel['text']
    D = destination22
    destinationLabel['text'] = D
destination_menu.add_command(label = "Main Place", command = setDestination22)

destination23 = "MicGinny's Bar & Grill"
def setDestination23():
    global DestinationLink
    DestinationLink = "https://www.google.com/maps/dir/Rochester+Institute+of+Technology,+1+Lomb+Memorial+Drive,+Rochester,+NY+14623/MicGinny's+Restaurant+%26+Sports+Pub,+2246+East+River+Road,+Rochester,+NY+14623/@43.0907736,-77.6865808,15z/data=!3m1!4b1!4m14!4m13!1m5!1m1!1s0x89d14c603a147e55:0xbe7eb31ed5e22c03!2m2!1d-77.6756115!2d43.0832027!1m5!1m1!1s0x89d14c932d6da297:0x15557b4a018c31a8!2m2!1d-77.6762844!2d43.0982536!3e1"
    D = destinationLabel['text']
    D = destination23
    destinationLabel['text'] = D
destination_menu.add_command(label = "MicGinny's Bar & Grill", command = setDestination23)

destination24 = "McKenzies"
def setDestination24():
    global DestinationLink
    DestinationLink = "https://www.google.com/maps/dir/Rochester+Institute+of+Technology,+1+Lomb+Memorial+Drive,+Rochester,+NY+14623/McKenzie's+Irish+Pub,+West+Henrietta+Road,+Rochester,+NY/@43.0826541,-77.6911663,13z/data=!3m1!4b1!4m14!4m13!1m5!1m1!1s0x89d14c603a147e55:0xbe7eb31ed5e22c03!2m2!1d-77.6756115!2d43.0832027!1m5!1m1!1s0x89d14b9d403a3521:0x73e99d62c899cb1c!2m2!1d-77.640877!2d43.0807788!3e1"
    D = destinationLabel['text']
    D = destination24
    destinationLabel['text'] = D
destination_menu.add_command(label = "McKenzies", command = setDestination24)

destination25 = "Nashville's"
def setDestination25():
    global DestinationLink
    DestinationLink = "https://www.google.com/maps/dir/Rochester+Institute+of+Technology,+1+Lomb+Memorial+Drive,+Rochester,+NY+14623/Nashvilles,+4853+West+Henrietta+Road,+Henrietta,+NY+14467/@43.071051,-77.705791,13z/data=!3m1!4b1!4m14!4m13!1m5!1m1!1s0x89d14c603a147e55:0xbe7eb31ed5e22c03!2m2!1d-77.6756115!2d43.0832027!1m5!1m1!1s0x89d14c081b8a6f6f:0x2b83c23f86753e58!2m2!1d-77.6542417!2d43.0556786!3e1"
    D = destinationLabel['text']
    D = destination25
    destinationLabel['text'] = D
destination_menu.add_command(label = "Nashville's", command = setDestination25)

destination26 = "Genessee Valley-Henrietta Moose"
def setDestination26():
    global DestinationLink
    DestinationLink = "https://www.google.com/maps/dir/Rochester+Institute+of+Technology,+1+Lomb+Memorial+Drive,+Rochester,+NY+14623/Genesee+Valley+Henrietta+Moose+Lodge+2290,+5375+West+Henrietta+Road,+West+Henrietta,+NY+14586/@43.0628499,-77.7444902,12z/data=!3m1!4b1!4m14!4m13!1m5!1m1!1s0x89d14c603a147e55:0xbe7eb31ed5e22c03!2m2!1d-77.6756115!2d43.0832027!1m5!1m1!1s0x89d14ea9de69c84b:0xe6ccd64fbf1629a8!2m2!1d-77.6595017!2d43.0455529!3e1"
    D = destinationLabel['text']
    D = destination26
    destinationLabel['text'] = D
destination_menu.add_command(label = "Genessee Valley-Henrietta Moose", command = setDestination26)

destination27 = "East Ridge Billiards"
def setDestination27():
    global DestinationLink
    DestinationLink = "https://www.google.com/maps/dir/Rochester+Institute+of+Technology,+1+Lomb+Memorial+Drive,+Rochester,+NY+14623/East+Ridge+Billiards,+529+East+Ridge+Road,+Rochester,+NY+14621/@43.1397743,-77.7100763,12z/data=!3m1!4b1!4m14!4m13!1m5!1m1!1s0x89d14c603a147e55:0xbe7eb31ed5e22c03!2m2!1d-77.6756115!2d43.0832027!1m5!1m1!1s0x89d6b67ef0757ead:0x74b7d7e18efd71d3!2m2!1d-77.6043411!2d43.1949773!3e1"
    D = destinationLabel['text']
    D = destination27
    destinationLabel['text'] = D
destination_menu.add_command(label = "East Ridge Billiards", command = setDestination27)

destination28 = "Classic Billiards"
def setDestination28():
    global DestinationLink
    DestinationLink = "https://www.google.com/maps/dir/Rochester+Institute+of+Technology,+1+Lomb+Memorial+Drive,+Rochester,+NY+14623/Camelot+Billards,+529+East+Ridge+Road,+Rochester,+NY+14621/@43.1397743,-77.7100763,12z/data=!3m1!4b1!4m14!4m13!1m5!1m1!1s0x89d14c603a147e55:0xbe7eb31ed5e22c03!2m2!1d-77.6756115!2d43.0832027!1m5!1m1!1s0x89d6b67ef070dd35:0xa0c8ee490bb3f013!2m2!1d-77.6043586!2d43.1949758!3e1"
    D = destinationLabel['text']
    D = destination28
    destinationLabel['text'] = D
destination_menu.add_command(label = "Classic Billiards", command = setDestination28)

destination29 = "6 Pockets Billiards"
def setDestination29():
    global DestinationLink
    DestinationLink = "https://www.google.com/maps/dir/Rochester+Institute+of+Technology,+1+Lomb+Memorial+Drive,+Rochester,+NY+14623/Six+Pockets,+Rochester,+NY+14621/@43.1354606,-77.7024478,12z/data=!3m1!4b1!4m14!4m13!1m5!1m1!1s0x89d14c603a147e55:0xbe7eb31ed5e22c03!2m2!1d-77.6756115!2d43.0832027!1m5!1m1!1s0x89d6b5d155555555:0x4bcfcfbb1dfe8f9e!2m2!1d-77.6021795!2d43.1874968!3e1"
    D = destinationLabel['text']
    D = destination29
    destinationLabel['text'] = D
destination_menu.add_command(label = "6 Pockets Billiards", command = setDestination29)

destination30 = "Camelot Billiards"
def setDestination30():
    global DestinationLink
    DestinationLink = "https://www.google.com/maps/dir/Rochester+Institute+of+Technology,+1+Lomb+Memorial+Drive,+Rochester,+NY+14623/Camelot+Billards,+529+East+Ridge+Road,+Rochester,+NY+14621/@43.1397743,-77.7100763,12z/data=!3m1!4b1!4m14!4m13!1m5!1m1!1s0x89d14c603a147e55:0xbe7eb31ed5e22c03!2m2!1d-77.6756115!2d43.0832027!1m5!1m1!1s0x89d6b67ef070dd35:0xa0c8ee490bb3f013!2m2!1d-77.6043586!2d43.1949758!3e1"
    D = destinationLabel['text']
    D = destination30
    destinationLabel['text'] = D
destination_menu.add_command(label = "Camelot Billiards", command = setDestination30)

menubar.add_cascade(label="Choose Destination", menu=destination_menu) # to create menu title


    ################################
    #                              #
    #        FILTER MENU           #
    #                              #
    ################################

filter_menu= Menu(menubar, tearoff=0)

def filterIsPub():
    ResetLabels()
    newObjList = []
    for Obj in InfoObjList:
        if Obj.Class == "Restaurant/Tavern":
            newObjList += [Obj]
    L=len(newObjList)
    for i in range(L):
        Lbl = LabelsList[i]['text']
        Lbl = newObjList[i].Name
        LabelsList[i]['text'] = Lbl
filter_menu.add_command(label = "Is tavern", command = filterIsPub)

def filterIsPoolHall():
    ResetLabels()
    newObjList = []
    for Obj in InfoObjList:
        if Obj.Class == "Pool hall":
            newObjList += [Obj]
    L=len(newObjList)
    for i in range(L):
        Lbl = LabelsList[i]['text']
        Lbl = newObjList[i].Name
        LabelsList[i]['text'] = Lbl
filter_menu.add_command(label = "Is pool hall", command = filterIsPoolHall)

def filterPizza():
    ResetLabels()
    newObjList = []
    for Obj in InfoObjList:
        if Obj.Pizza == "Yes":
            newObjList += [Obj]
    L=len(newObjList)
    for i in range(L):
        Lbl = LabelsList[i]['text']
        Lbl = newObjList[i].Name
        LabelsList[i]['text'] = Lbl
filter_menu.add_command(label = "Has pizza", command = filterPizza)

def filterBurgers():
    ResetLabels()
    newObjList = []
    for Obj in InfoObjList:
        if Obj.Burgers == "Yes":
            newObjList += [Obj]
    L=len(newObjList)
    for i in range(L):
        Lbl = LabelsList[i]['text']
        Lbl = newObjList[i].Name
        LabelsList[i]['text'] = Lbl
filter_menu.add_command(label = "Has burgers", command = filterBurgers)

def filterSandwiches():
    ResetLabels()
    newObjList = []
    for Obj in InfoObjList:
        if Obj.Sandwiches == "Yes":
            newObjList += [Obj]
    L=len(newObjList)
    for i in range(L):
        Lbl = LabelsList[i]['text']
        Lbl = newObjList[i].Name
        LabelsList[i]['text'] = Lbl
filter_menu.add_command(label = "Is tavern", command = filterSandwiches)

menubar.add_cascade(label="Filter", menu=filter_menu) # to create menu title

    ################################
    #                              #
    #          SORT MENU           #
    #                              #
    ################################

sort_menu= Menu(menubar, tearoff=0)

def sortByDistanceAscending():
    InfoObjList2 = InfoObjList[:]
    L=len(InfoObjList2)
    for i in range (L):
        for j in range(i,L):
            if InfoObjList2[j].Distance < InfoObjList2[i].Distance:
                InfoObjList2[i],InfoObjList2[j]=InfoObjList2[j], InfoObjList2[i] # to swap elements
    for i in range(L):
        D = LabelsList[i]['text']
        D = (InfoObjList2[i].Name + ": "+ str(InfoObjList2[i].Distance))
        LabelsList[i]['text'] = D
sort_menu.add_command(label = "Distance-ascending", command = sortByDistanceAscending)

def sortByDistanceDescending():
    InfoObjList2 = InfoObjList[:]
    L=len(InfoObjList2)
    for i in range (L):
        for j in range(i,L):
            if InfoObjList2[j].Distance > InfoObjList2[i].Distance:
                InfoObjList2[i],InfoObjList2[j]=InfoObjList2[j], InfoObjList2[i] # to swap elements
    for i in range(L):
        D = LabelsList[i]['text']
        D = (InfoObjList2[i].Name + ": "+ str(InfoObjList2[i].Distance))
        LabelsList[i]['text'] = D
sort_menu.add_command(label = "Distance-descending", command = sortByDistanceDescending)

def sortByPriceAscending():
    InfoObjList2 = InfoObjList[:]
    L=len(InfoObjList2)
    for i in range (L):
        for j in range(i,L):
            if InfoObjList2[j].Distance < InfoObjList2[i].Distance:
                InfoObjList2[i],InfoObjList2[j]=InfoObjList2[j], InfoObjList2[i] # to swap elements
    for i in range(L):
        D = LabelsList[i]['text']
        D = (InfoObjList2[i].Name)
        LabelsList[i]['text'] = D
sort_menu.add_command(label = "Price-descending", command = sortByPriceAscending)

def sortByPriceDescending():
    InfoObjList2 = InfoObjList[:]
    L=len(InfoObjList2)
    for i in range (L):
        for j in range(i,L):
            if InfoObjList2[j].Price > InfoObjList2[i].Price:
                InfoObjList2[i],InfoObjList2[j]=InfoObjList2[j], InfoObjList2[i] # to swap elements
    for i in range(L):
        D = LabelsList[i]['text']
        D = (InfoObjList2[i].Name)
        LabelsList[i]['text'] = D
sort_menu.add_command(label = "Price-ascending", command = sortByPriceDescending)

def sortByTablesAscending():
    InfoObjList2 = InfoObjList[:]
    L=len(InfoObjList2)
    for i in range (L):
        for j in range(i,L):
            if InfoObjList2[j].NumOfTables < InfoObjList2[i].NumOfTables:
                InfoObjList2[i],InfoObjList2[j]=InfoObjList2[j], InfoObjList2[i] # to swap elements
    for i in range(L):
        D = LabelsList[i]['text']
        D = (InfoObjList2[i].Name + ": "+ str(InfoObjList2[i].NumOfTables))
        LabelsList[i]['text'] = D
sort_menu.add_command(label = "Num of tables-ascending", command = sortByTablesAscending)

def sortByTablesDescending():
    InfoObjList2 = InfoObjList[:]
    L=len(InfoObjList2)
    for i in range (L):
        for j in range(i,L):
            if InfoObjList2[j].NumOfTables > InfoObjList2[i].NumOfTables:
                InfoObjList2[i],InfoObjList2[j]=InfoObjList2[j], InfoObjList2[i] # to swap elements
    for i in range(L):
        D = LabelsList[i]['text']
        D = (InfoObjList2[i].Name + ": "+ str(InfoObjList2[i].NumOfTables))
        LabelsList[i]['text'] = D
sort_menu.add_command(label = "Num of tables-descending", command = sortByTablesDescending)

menubar.add_cascade(label="Sort", menu=sort_menu) # to create menu title

    ################################
    #                              #
    #          INFO MENU           #
    #                              #
    ################################


def MessageDisp():
    master = Tk()
    master.title("App Info")
    TEXT = "Created by \n Daniel Roy Barman \n for Map Project"
    msg = Message(master, text = TEXT)
    msg.config(bg='white', font=('times', 20, 'italic'))
    msg.pack( )
        
infomenu = Menu(menubar, tearoff=0)
infomenu.add_command(label="About this app", command=MessageDisp)

menubar.add_cascade(label="Info", menu=infomenu) # to create menu title


    ################################
    #                              #
    #       CONTACT ME MENU        #
    #                              #
    ################################

contactMEmenu = Menu(menubar, tearoff=0)

def Contact():
    subRoot =Tk()
    subRoot.title('Contact the developer')
    canvas2 = Canvas(subRoot, width=200, height=100)
    canvas2.pack()
    Label(subRoot, text="First Name").place(relx=0.05, rely=0.06)
    Label(subRoot, text="Last Name").place(relx=0.05, rely=0.29)
    Label(subRoot, text="Message Text").place(relx=0.05, rely=0.55)

    e1 = Entry(subRoot)
    e2 = Entry(subRoot)
    e3 = Entry(subRoot)

    e1.place(relx=0.4, rely=0.06)
    e2.place(relx=0.4, rely=0.29)
    e3.place(relx=0.4, rely=0.55)
    
    def sendMail():
        smtpObj = smtplib.SMTP("smtp.mail.yahoo.com", 587)
        smtpObj.ehlo()
        smtpObj.starttls()
        smtpObj.login("danielrb11@yahoo.com", "lordsuperdanbrucelee1997")
        smtpObj.sendmail("danielrb11@yahoo.com", "daniel.rb4@gmail.com",str(e3.get())+"\n\n"+str(e1.get())+" "+str(e2.get()))

    b6 = Button(subRoot, text=" send", command=sendMail, anchor='w')
    b6.place(relx=0.9, rely=0.4)
    b6.config(height=1, width=6) # button size
    b6.config(font=('calibri', 15)) # font
    b6.config(bg="blue", fg="white", activebackground="black", activeforeground="white")
    b6.config(highlightbackground="grey")
    b6.pack()

    subRoot.mainloop()
    
contactMEmenu.add_command(label="Send Email", command=Contact)

menubar.add_cascade(label="Contact the developer", menu=contactMEmenu) # to create menu title

    ################################
    #                              #
    #          QUIT MENU           #
    #                              #
    ################################

    
quitmenu = Menu(menubar, tearoff=0)
quitmenu.add_command(label="Are you sure?", command=root.quit)
menubar.add_cascade(label="Quit", menu=quitmenu) # to create menu title

root.config(menu=menubar)
root.mainloop()

